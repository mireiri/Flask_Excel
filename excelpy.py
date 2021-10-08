import openpyxl
from openpyxl.chart import BarChart, Reference


def excelpy(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # 定刻よりも早く出発（到着）していれば◯、していなければ✕を表す列を追加する
    STD = []
    for col in ws.iter_cols(min_row=2, min_col=4, max_col=4):
        for cell in col:
            STD.append(cell.value)

    ATD = []
    for col in ws.iter_cols(min_row=2, min_col=6, max_col=6):
        for cell in col:
            ATD.append(cell.value)
        
    DEP = []
    for i, k in zip(STD, ATD):
        if i >= k:  
            DEP.append('◯')
        else:
            DEP.append('✕')

    ws['I1'] = '定時出発'
    for col in ws.iter_cols(min_row=2, min_col=9, max_col=9):
        for cell, i in zip(col, DEP):
            cell.value = i

    STA = []
    for col in ws.iter_cols(min_row=2, min_col=5, max_col=5):
        for cell in col:
            STA.append(cell.value)

    ATA = []
    for col in ws.iter_cols(min_row=2, min_col=7, max_col=7):
        for cell in col:
            ATA.append(cell.value)

    ARR = []
    for i, k in zip(STA, ATA):
        if i >= k:
            ARR.append('◯')
        else:
            ARR.append('✕')

    ws['J1'] = '定時到着'
    for col in ws.iter_cols(min_row=2, min_col=10, max_col=10):
        for cell, i in zip(col, ARR):
            cell.value = i

    # 乗車率を表す列を追加する
    ws['K1'] = '乗車率'
    LOAD_FACTOR = []
    for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
        for cell in col:
            lf = cell.value / 45
            LOAD_FACTOR.append('{:.0%}'.format(lf))
    
    for col in ws.iter_cols(min_row=2, min_col=11, 
                            max_row=len(LOAD_FACTOR)+1, max_col=11):
        for cell, i in zip(col, LOAD_FACTOR):
            cell.value = i
    
    # 区間ごとの運行数を反映する
    route_num = {}
    for col in ws.iter_cols(min_row=2, min_col=3, max_col=3):
        for cell in col:
            if cell.value in route_num:
                route_num[cell.value] += 1
            else:
                route_num[cell.value] = 1

    route = [i for i in route_num.keys()]
    route.insert(0, '区間')

    num = [i for i in route_num.values()]
    num.insert(0, '運行数')

    for col in ws.iter_cols(min_row=1, min_col=12, 
                            max_row=len(route), max_col=12):
        for cell, r in zip(col, route):
            cell.value = r

    for col in ws.iter_cols(min_row=1, min_col=13, 
                            max_row=len(num), max_col=13):
        for cell, n in zip(col, num):
            cell.value = n

    # 区間ごとの運行数をグラフにする
    chart = BarChart()

    plot_value = Reference(ws, min_row=1, min_col=13, max_row=5, max_col=13)
    chart.add_data(plot_value, titles_from_data=True)

    x = Reference(ws, min_row=2, min_col=12, max_row=5, max_col=12) 
    chart.set_categories(x)
    ws.add_chart(chart, 'L7')
    
    # 定時出発率、定時到着率、平均乗車率を反映する
    ws['N1'] = '定時出発率'
    dep_result = []
    for col in ws.iter_cols(min_row=2, min_col=9, max_col=9):
        for cell in col:
            dep_result.append(cell.value)

    dep_rate = dep_result.count('◯') / len(dep_result)
    ws['N2'] = '{:.0%}'.format(dep_rate)

    ws['O1'] = '定時到着率'
    arr_result = []
    for col in ws.iter_cols(min_row=2, min_col=10, max_col=10):
        for cell in col:
            arr_result.append(cell.value)

    arr_rate = arr_result.count('◯') / len(arr_result)
    ws['O2'] = '{:.0%}'.format(arr_rate)

    ws['P1'] = '平均乗車率'
    lf_result = 0
    for col in ws.iter_cols(min_row=2, min_col=8, max_col=8):
        for cell in col:
            lf_result += cell.value / 45
    
    lf_result = lf_result / len(arr_result)
    ws['P2'] = '{:.0%}'.format(lf_result)

    # 編集結果を上書き保存してファイルパスを返す    
    wb.save(file_path)
    return file_path


# このファイルの動作確認
if __name__ == '__main__':
    result_file_path = excelpy('static/sample.xlsx')
    print(result_file_path)
